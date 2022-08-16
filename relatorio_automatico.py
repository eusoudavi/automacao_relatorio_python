import openpyxl
import graficos
from openpyxl.styles import Alignment, Border, Side


def rel_geral():
    nome_do_arquivo = 'teste.xlsx'
    planilha = openpyxl.load_workbook(nome_do_arquivo)

    aba_pesquisada = 'AGOSTO'
    aba_pesquisada = aba_pesquisada.upper().strip()

    # ----------------CRIAÇÃO DA ABA (WORKSHEETS)----------------
    if f'RELATÓRIO_{"GERAL"}' in planilha.sheetnames:  # Se o worksheet já existe, o programa seleciona a aba
        ws = planilha[f'RELATÓRIO_{"GERAL"}']
    else:  # Se não existir, cria o worksheet
        modelo = planilha[f'Modelo']
        ws = planilha.copy_worksheet(modelo)  # É criado com base no modelo na planilha
        ws.title = f'RELATÓRIO_{"GERAL"}'

    jornal_imp = {'semanas': [0, 0, 0, 0, 0], "nome": 'JORNAL IMPRESSO', "total": 0}
    jornal_dig = {'semanas': [0, 0, 0, 0, 0], "nome": 'JORNAL DIGITAL', "total": 0}
    card_dig = {'semanas': [0, 0, 0, 0, 0], "nome": 'CARD DIGITAL', "total": 0}
    carrosel = {'semanas': [0, 0, 0, 0, 0], "nome": 'CARROSSEL', "total": 0}
    tv = {'semanas': [0, 0, 0, 0, 0], "nome": 'TV', "total": 0}
    radio = {'semanas': [0, 0, 0, 0, 0], "nome": 'RÁDIO', "total": 0}
    carro_som = {'semanas': [0, 0, 0, 0, 0], "nome": 'CARRO DE SOM', "total": 0}
    mercham = {'semanas': [0, 0, 0, 0, 0], "nome": 'MERCHAN', "total": 0}

    midias = [jornal_imp, jornal_dig, card_dig, carrosel, tv, radio, carro_som, mercham]

    for celulas in ws["B"]:      # Varrer a primeira coluna para encontrar o nome do consultor.
        if type(celulas.value) == str:      # Verificar se é uma "string" para eliminar as células vazias
            if 'Nome do Consultor' in celulas.value:
                ws[f'B{celulas.row}'].value = ws.title.capitalize()

    for abas in planilha.worksheets:                # Varrer as abas da planilha e
        if aba_pesquisada in abas.title.upper():    # verificar se há uma aba onde os dados serão consultados.
            for celulas in abas["A"]:               # Varrer a primeira coluna para encontrar o nome do consultor.
                if type(celulas.value) == str:      # Verificar se é uma "string" para eliminar as células vazias
                    coluna_b = abas[f'B{celulas.row}'].value
                    coluna_c = abas[f'C{celulas.row}'].value
                    for i in range(5):          # Alimentação do Discionário com as informações da planilha
                        if f'SEMANA 0{i+1}' in coluna_b and jornal_imp['nome'] in coluna_c:
                            jornal_imp['semanas'][i] += 1
                            jornal_imp['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and jornal_dig['nome'] in coluna_c:
                            jornal_dig['semanas'][i] += 1
                            jornal_dig['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and card_dig['nome'] in coluna_c:
                            card_dig['semanas'][i] += 1
                            card_dig['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and carrosel['nome'] in coluna_c:
                            carrosel['semanas'][i] += 1
                            carrosel['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and tv['nome'] in coluna_c:
                            tv['semanas'][i] += 1
                            tv['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and radio['nome'] in coluna_c:
                            radio['semanas'][i] += 1
                            radio['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and carro_som['nome'] in coluna_c:
                            carro_som['semanas'][i] += 1
                            carro_som['total'] += 1
                        if f'SEMANA 0{i+1}' in coluna_b and mercham['nome'] in coluna_c:
                            mercham['semanas'][i] += 1
                            mercham['total'] += 1
    # Plotar no relatório os resultados
    for abas in planilha.worksheets:
        if ws.title in abas.title.upper():
            for celulas in abas["B"]:
                if type(celulas.value) == str:
                    for i in range(len(midias)):
                        if celulas.value.upper().strip() == midias[i]['nome']:
                            abas[f'C{celulas.row}'].value = midias[i]['semanas'][0]
                            abas[f'D{celulas.row}'].value = midias[i]['semanas'][1]
                            abas[f'E{celulas.row}'].value = midias[i]['semanas'][2]
                            abas[f'F{celulas.row}'].value = midias[i]['semanas'][3]
                            abas[f'G{celulas.row}'].value = midias[i]['semanas'][4]
    # Plotar no Relatório os valores totais das ações
    borda_grossa = Side(border_style='medium', color='000000')
    borda_fina = Side(border_style='hair', color='000000')

    ws['H3'].value = 'TOTAL'
    ws['H3'].alignment = Alignment(horizontal='center')     # Ajustar o texto para ficar centralizado na celula
    # Vamos criar bordas para os valores criados
    ws['H3'].border = Border(top=borda_grossa, right=borda_grossa, bottom=borda_grossa, left=borda_grossa)
    for i in range(len(midias)):
        ws[f'H{i + 4}'].value = midias[i]['total']
        # Criar linhas finas no meio da tabela e grossas ao redor
        if i == len(midias) - 1:
            ws[f'H{i + 4}'].border = Border(left=borda_grossa, right=borda_grossa, bottom=borda_grossa)
        else:
            ws[f'H{i + 4}'].border = Border(left=borda_grossa, right=borda_grossa, bottom=borda_fina)

    planilha.save('teste.xlsx')
    graficos.AutoChart(f'RELATÓRIO_GERAL', nome_do_arquivo)
    graficos.grafico_pizza(f'RELATÓRIO_GERAL', nome_do_arquivo)


def rel_individual(name):
    """
    Função executa uma rotina automática de leitura e criação de relatórios.
    :param name:
    :return:
    """
    # planilha = openpyxl.load_workbook('HUB_CalendárioAgosto_v5.xlsx')
    nome_do_arquivo = 'teste.xlsx'
    planilha = openpyxl.load_workbook(nome_do_arquivo)

    aba_pesquisada = 'AGOSTO'
    aba_pesquisada = aba_pesquisada.upper().strip()
    # CRIAR ROTINA PARA CAPTURAR O MÊS ATUAL E PERGUNTAR AO USUÁRIO SE ELE DESEJA FAZER O RELATÓRIO ATUALIZADO DO MÊS

    nome_do_consultor = str(name)           # Nomes dos consultores
    nome_do_consultor = nome_do_consultor.upper().strip()

    jornal_imp = {'semanas': [0, 0, 0, 0, 0], "nome": 'JORNAL IMPRESSO'}
    jornal_dig = {'semanas': [0, 0, 0, 0, 0], "nome": 'JORNAL DIGITAL'}
    card_dig = {'semanas': [0, 0, 0, 0, 0], "nome": 'CARD DIGITAL'}
    carrosel = {'semanas': [0, 0, 0, 0, 0], "nome": 'CARROSSEL'}
    tv = {'semanas': [0, 0, 0, 0, 0], "nome": 'TV'}
    radio = {'semanas': [0, 0, 0, 0, 0], "nome": 'RÁDIO'}
    carro_som = {'semanas': [0, 0, 0, 0, 0], "nome": 'CARRO DE SOM'}
    mercham = {'semanas': [0, 0, 0, 0, 0], "nome": 'MERCHAN'}

    midias = [jornal_imp, jornal_dig, card_dig, carrosel, tv, radio, carro_som, mercham]

    # ----------------CRIAÇÃO DAS ABAS (WORKSHEETS)----------------
    if f'RELATÓRIO_{nome_do_consultor}' in planilha.sheetnames:  # Se o worksheet já existe, o programa seleciona a aba
        aba_relatorio = planilha[f'RELATÓRIO_{nome_do_consultor}']
    else:  # Se não existir, cria o worksheet
        modelo = planilha[f'Modelo']
        aba_relatorio = planilha.copy_worksheet(modelo)  # É criado com base no modelo na planilha
        aba_relatorio.title = f'RELATÓRIO_{nome_do_consultor}'

    # ----------------PROCESSAMENTO DO RELATÓRIO----------------
    # Colocar o nome do consultor no começo da planilha
    for celulas in aba_relatorio["B"]:      # Varrer a primeira coluna para encontrar o nome do consultor.
        if type(celulas.value) == str:      # Verificar se é uma "string" para eliminar as células vazias
            if 'Nome do Consultor' in celulas.value:
                aba_relatorio[f'B{celulas.row}'].value = nome_do_consultor.capitalize()

    # Varrer as abas da planilha para encontrar a aba onde deseja obter os dados
    for abas in planilha.worksheets:                # Varrer as abas da planilha e
        if aba_pesquisada in abas.title.upper():    # verificar se há uma aba onde os dados serão consultados.
            for celulas in abas["A"]:               # Varrer a primeira coluna para encontrar o nome do consultor.
                if type(celulas.value) == str:      # Verificar se é uma "string" para eliminar as células vazias
                    if nome_do_consultor in celulas.value.upper():      # IMPORTANTE
                        for i in range(5):          # Alimentação do Discionário com as informações da planilha
                            coluna_b = abas[f'B{celulas.row}'].value
                            coluna_c = abas[f'C{celulas.row}'].value
                            if f'SEMANA 0{i+1}' in coluna_b and jornal_imp['nome'] in coluna_c:
                                jornal_imp['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and jornal_dig['nome'] in coluna_c:
                                jornal_dig['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and card_dig['nome'] in coluna_c:
                                card_dig['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and carrosel['nome'] in coluna_c:
                                carrosel['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and tv['nome'] in coluna_c:
                                tv['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and radio['nome'] in coluna_c:
                                radio['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and carro_som['nome'] in coluna_c:
                                carro_som['semanas'][i] += 1
                            if f'SEMANA 0{i+1}' in coluna_b and mercham['nome'] in coluna_c:
                                mercham['semanas'][i] += 1
    # Varrer as abas agora para encontrar onde o relatório foi criado para plotar os dados do Discionário
    for abas in planilha.worksheets:
        if aba_relatorio.title in abas.title.upper():
            for celulas in abas["B"]:
                if type(celulas.value) == str:
                    for i in range(len(midias)):
                        if celulas.value.upper().strip() == midias[i]['nome']:
                            abas[f'C{celulas.row}'].value = midias[i]['semanas'][0]
                            abas[f'D{celulas.row}'].value = midias[i]['semanas'][1]
                            abas[f'E{celulas.row}'].value = midias[i]['semanas'][2]
                            abas[f'F{celulas.row}'].value = midias[i]['semanas'][3]
                            abas[f'G{celulas.row}'].value = midias[i]['semanas'][4]

    planilha.save('teste.xlsx')
    graficos.AutoChart(f'RELATÓRIO_{nome_do_consultor}', nome_do_arquivo)
    # graficos.colorir(f'RELATÓRIO_SUELEN', 'teste.xlsx')
    return 'SALVO'


# rel_individual('Suelen')
