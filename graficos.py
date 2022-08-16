import openpyxl
from openpyxl.chart import BarChart, Reference, PieChart, ProjectedPieChart
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors


def colorir(aba_pesquisada, arquivo):
    planilha = openpyxl.load_workbook(arquivo)
    ws = planilha[f'{aba_pesquisada}']

    red_fill = PatternFill(start_color='FFFF0000',
                           end_color='FFFF0000',
                           fill_type='solid')

    ws['B2'].fill = red_fill
    planilha.save(arquivo)


def AutoChart(aba_pesquisada, arquivo):
    """
    Função para elaboração de graficos no excel
    :param aba_pesquisada: WORKSHEET FROM EXCEL WITH DATA FOR CHART
    :param arquivo: WORKBOOK
    :return:
    """
    # A função irá abrir o arquivo definido por quem está fazendo o relatório
    planilha = openpyxl.load_workbook(arquivo)
    for index, abas in enumerate(planilha.worksheets):
        if aba_pesquisada.upper().strip() in abas.title.upper():
            ws = planilha.worksheets[index]
            # categoria = Reference(ws, min_col=1, min_row=2, max_row=9)
            categoria = Reference(ws, min_row=3, min_col=3, max_col=7)
            # valores = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=9)
            valores = Reference(ws, min_col=2, max_col=7, min_row=3, max_row=11)

            graphic = BarChart()

            graphic.add_data(valores, titles_from_data=True, from_rows=True)
            graphic.set_categories(categoria)

            ws.add_chart(graphic, 'B13')

            # print(valores)
            planilha.save(arquivo)
            print(f'Gráfico {aba_pesquisada} feito')
            return 'Salvo'


# AutoChart('RELATÓRIO_VANDERSON', 'teste.xlsx')


def grafico_pizza(aba_pesquisada, arquivo):
    planilha = openpyxl.load_workbook(arquivo)

    for index, abas in enumerate(planilha.worksheets):
        if aba_pesquisada.upper().strip() in abas.title.upper():
            ws = planilha.worksheets[index]

            categoria = Reference(ws, min_col=2, min_row=4, max_row=11)
            valores = Reference(ws, min_col=8, min_row=4, max_row=11)

            pizza = PieChart()

            pizza.add_data(valores, titles_from_data=True)
            pizza.set_categories(categoria)
            pizza.title = "Distribuição Geral do Mês"

            ws.add_chart(pizza, 'G13')

            # print(valores)
            planilha.save(arquivo)
            print(f'Gráfico Pizza {aba_pesquisada} feito')
            return 'Salvo'

